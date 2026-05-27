"""
Extractor automatico de morosos - Comunidad Feliz
Corre via GitHub Actions cada noche
"""
import time, datetime, os, shutil, re, json, base64
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# CONFIGURACION
EMAIL        = os.environ['CF_EMAIL']
PASSWORD     = os.environ['CF_PASSWORD']
FOLDER_ID    = '17vrvDsPfunupOoL6i2Mc-Xsv7IYl5wOS'
FECHA_ACTUAL = datetime.datetime.now().strftime('%Y-%m')
MESES_CORTE  = 3
WORK_DIR     = '/tmp/morosos'
os.makedirs(WORK_DIR, exist_ok=True)

# GOOGLE DRIVE AUTH
def get_drive_service():
    creds_json = os.environ['GOOGLE_CREDENTIALS_JSON']
    creds_dict = json.loads(creds_json)
    creds = Credentials.from_service_account_info(
        creds_dict,
        scopes=['https://www.googleapis.com/auth/drive']
    )
    return build('drive', 'v3', credentials=creds)

def subir_a_drive(service, ruta_local, nombre_archivo, folder_id):
    # ID fijo del reporte principal que ya existe en Drive
    REPORTE_FILE_ID = '1WHP32-gkVLhJ1g3VgWyeLLKSQbo0MMEp'

    media = MediaFileUpload(ruta_local,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    if 'REPORTE_MOROSOS' in nombre_archivo:
        # Siempre actualizar el archivo existente (nunca crear nuevo)
        service.files().update(
            fileId=REPORTE_FILE_ID,
            body={'name': nombre_archivo},
            media_body=media
        ).execute()
        print(f'  Actualizado en Drive: {nombre_archivo}')
    else:
        # Para los Excel individuales, buscar si existe y actualizar
        query = f"name='{nombre_archivo}' and '{folder_id}' in parents and trashed=false"
        results = service.files().list(q=query, fields='files(id,name)').execute()
        files = results.get('files', [])
        if files:
            service.files().update(
                fileId=files[0]['id'],
                media_body=media
            ).execute()
            print(f'  Actualizado: {nombre_archivo}')
        else:
            print(f'  Omitido (no existe en Drive): {nombre_archivo}')

# SELENIUM CHROME
def iniciar_chrome():
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    options.add_experimental_option('prefs', {
        'download.default_directory': WORK_DIR,
        'download.prompt_for_download': False,
        'download.directory_upgrade': True,
        'safebrowsing.enabled': True
    })
    driver = webdriver.Chrome(options=options)
    driver.execute_cdp_cmd('Page.setDownloadBehavior',
        {'behavior': 'allow', 'downloadPath': WORK_DIR})
    return driver

def click(driver, el):
    driver.execute_script('arguments[0].scrollIntoView(true);', el)
    driver.execute_script('arguments[0].click();', el)

def limpiar_crdownload():
    for f in os.listdir(WORK_DIR):
        if f.endswith('.crdownload'):
            try:
                os.remove(os.path.join(WORK_DIR, f))
            except:
                pass

def esperar_xlsx(timeout=60):
    for _ in range(timeout):
        files = [f for f in os.listdir(WORK_DIR)
                 if f.endswith('.xlsx') and not f.endswith('.crdownload')]
        if files:
            return max([os.path.join(WORK_DIR, f) for f in files], key=os.path.getctime)
        time.sleep(1)
    return None

def limpiar_xlsx_sin_nombre():
    for f in os.listdir(WORK_DIR):
        if f.endswith('.xlsx') and not f.startswith(FECHA_ACTUAL):
            try:
                os.remove(os.path.join(WORK_DIR, f))
            except:
                pass

# DESCARGA DE MOROSOS
def descargar_morosos():
    driver = iniciar_chrome()
    wait   = WebDriverWait(driver, 30)
    archivos_descargados = []

    try:
        print('Iniciando sesion...')
        driver.get('https://app.comunidadfeliz.com')
        wait.until(EC.presence_of_element_located((By.NAME, 'email'))).send_keys(EMAIL)
        driver.find_element(By.NAME, 'password').send_keys(PASSWORD)
        click(driver, driver.find_element(By.XPATH, "//button[@type='submit']"))
        time.sleep(8)
        print(f'Login OK: {driver.current_url}')

        driver.get('https://app.comunidadfeliz.com/mis_comunidades')
        time.sleep(8)
        for _ in range(15):
            prev = len(driver.find_elements(By.XPATH, "//a[contains(@href,'/comunidades/')]"))
            driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
            time.sleep(2)
            curr = len(driver.find_elements(By.XPATH, "//a[contains(@href,'/comunidades/')]"))
            if curr == prev and curr > 0:
                break

        links_raw = driver.find_elements(By.XPATH, "//a[contains(@href,'/comunidades/')]")
        edificios = []
        seen = set()
        excluir = ['autoinstall','base_plans','academy','usuarios','cerrar','as_manager','gamification']
        for l in links_raw:
            href  = l.get_attribute('href') or ''
            texto = l.text.strip().replace('\n',' ').strip()
            texto = re.split(r'RUT|Ultimo|Emitido|\d{2}\.\d{3}', texto)[0].strip()
            if href and texto and len(texto) > 4 and href not in seen:
                if not any(x in href for x in excluir):
                    seen.add(href)
                    edificios.append({'href': href, 'nombre': texto})

        print(f'{len(edificios)} edificio(s) encontrado(s)')

        for edificio in edificios:
            nombre = edificio['nombre']
            print(f'\nProcesando: {nombre}')
            try:
                # Solo limpiar archivos temporales incompletos
                limpiar_crdownload()

                driver.get(edificio['href'])
                time.sleep(6)
                driver.get('https://app.comunidadfeliz.com/boletas/morosity')
                time.sleep(5)

                if 'morosity' not in driver.current_url:
                    print('  Sin morosidad')
                    continue

                btn = driver.execute_script(
                    "return document.getElementById('dropdown_morosity_menu_options');")
                if not btn:
                    print('  Sin morosos')
                    continue

                resultado = driver.execute_script("""
                    var btn = document.getElementById('dropdown_morosity_menu_options');
                    btn.click();
                    var parent = btn.parentElement;
                    var links = parent.querySelectorAll('a');
                    for (var i = 0; i < links.length; i++) {
                        if (links[i].innerHTML.toLowerCase().includes('excel') ||
                            links[i].innerText.toLowerCase().includes('excel')) {
                            links[i].click();
                            return 'OK: ' + links[i].innerText.trim();
                        }
                    }
                    return 'ERROR: Excel no encontrado';
                """)
                print(f'  {resultado}')
                time.sleep(3)

                # Esperar que se descargue y renombrar
                archivo = esperar_xlsx(60)
                if archivo:
                    nombre_limpio = re.sub(r'[<>:"/\\|?*]', '', nombre).strip()
                    nombre_final  = f'{FECHA_ACTUAL}_{nombre_limpio}.xlsx'
                    destino       = os.path.join(WORK_DIR, nombre_final)
                    shutil.copy2(archivo, destino)
                    os.remove(archivo)
                    archivos_descargados.append({
                        'nombre': nombre,
                        'ruta': destino,
                        'archivo': nombre_final
                    })
                    print(f'  Descargado: {nombre_final}')
                else:
                    print('  Sin descarga')

            except Exception as e:
                print(f'  Error: {e}')

    finally:
        driver.quit()

    return archivos_descargados

# ANALISIS DE MOROSOS
def analizar_morosos(archivos):
    meses_re = re.compile(
        r'(enero|febrero|marzo|abril|mayo|junio|julio|agosto|'
        r'septiembre|octubre|noviembre|diciembre|'
        r'ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)',
        re.IGNORECASE)

    def limpiar_num(v):
        if pd.isna(v): return 0
        s = str(v).replace('$','').replace('.','').replace(',','.').replace(' ','').strip()
        try: return float(s)
        except: return 0

    todos = []
    for item in archivos:
        edificio = item['nombre']
        ruta     = item['ruta']
        print(f'\nAnalizando: {edificio}')
        try:
            if not os.path.exists(ruta):
                print(f'  Archivo no encontrado: {ruta}')
                continue

            xls  = pd.ExcelFile(ruta)
            hoja = next((h for h in xls.sheet_names if 'unidad' in h.lower()), xls.sheet_names[0])

            df_raw = pd.read_excel(xls, sheet_name=hoja, header=None)
            fila_header = None
            for i, row in df_raw.iterrows():
                vals = [str(v).strip().lower() for v in row.values]
                if any('unidad' in v for v in vals) and any('total' in v for v in vals):
                    fila_header = i
                    break
            if fila_header is None:
                for i, row in df_raw.iterrows():
                    vals = [str(v).strip().lower() for v in row.values]
                    if any('unidad' in v for v in vals):
                        fila_header = i
                        break
            if fila_header is None:
                print('  No se encontro fila de encabezados')
                continue

            df = pd.read_excel(xls, sheet_name=hoja, header=fila_header)
            df.columns = df.columns.astype(str).str.strip()
            df = df.loc[:, df.columns.notna()]
            df = df.loc[:, ~df.columns.str.lower().isin(['nan','none',''])]
            df = df.dropna(how='all')

            col_unidad    = next((c for c in df.columns if 'unidad' in c.lower()), None)
            col_residente = next((c for c in df.columns if 'residente' in c.lower()), None)
            col_total     = next((c for c in df.columns
                                  if 'total' in c.lower() and 'deuda' in c.lower()), None)
            if not col_total:
                col_total = next((c for c in df.columns if 'total' in c.lower()), None)
            if not col_unidad:
                print('  No se encontro columna Unidad')
                continue

            idx_total  = df.columns.tolist().index(col_total) if col_total else len(df.columns)
            cols_meses = [c for c in df.columns.tolist()[:idx_total]
                          if meses_re.search(str(c))]

            for _, row in df.iterrows():
                unidad = str(row.get(col_unidad, '')).strip()
                if not unidad or unidad.lower() in ['nan','unidad','total','subtotal','']:
                    continue
                deuda_por_meses = sum(limpiar_num(row.get(c, 0)) for c in cols_meses)
                deuda_col_total = limpiar_num(row.get(col_total, 0)) if col_total else 0
                deuda_total = max(deuda_por_meses, deuda_col_total)
                if deuda_total <= 0:
                    continue
                meses_con_deuda = [c for c in cols_meses if limpiar_num(row.get(c, 0)) > 0]
                n_meses = len(meses_con_deuda)
                residente = str(row.get(col_residente, 'S/I')).strip() if col_residente else 'S/I'
                if n_meses > MESES_CORTE:
                    fila = {
                        'Edificio':        edificio,
                        'Unidad':          unidad,
                        'Residente':       residente,
                        'Deuda Total ($)': deuda_total,
                        'Meses en Deuda':  n_meses,
                    }
                    for cm in cols_meses:
                        v = limpiar_num(row.get(cm, 0))
                        if v > 0:
                            fila[cm] = v
                    todos.append(fila)

            n = sum(1 for t in todos if t['Edificio'] == edificio)
            print(f'  Morosos +{MESES_CORTE} meses: {n}')

        except Exception as e:
            print(f'  Error: {e}')

    return todos

# GENERAR EXCEL REPORTE
def generar_reporte(todos):
    if not todos:
        print('Sin morosos')
        return None

    df_out    = pd.DataFrame(todos).sort_values(['Edificio','Meses en Deuda'], ascending=[True,False])
    fecha_hoy = datetime.datetime.now().strftime('%Y-%m-%d')
    ruta_rep  = os.path.join(WORK_DIR, f'REPORTE_MOROSOS_+3meses_{fecha_hoy}.xlsx')

    with pd.ExcelWriter(ruta_rep, engine='openpyxl') as w:
        df_out.to_excel(w, sheet_name='Morosos +3 meses', index=False)
        res = df_out.groupby('Edificio').agg(
            Unidades=('Unidad','count'),
            Deuda_Total=('Deuda Total ($)','sum'),
            Meses_Promedio=('Meses en Deuda','mean')
        ).round(1).reset_index()
        res['Deuda_Total'] = res['Deuda_Total'].apply(lambda x: f'${x:,.0f}')
        res.to_excel(w, sheet_name='Resumen por Edificio', index=False)

    wb = load_workbook(ruta_rep)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for cell in ws[1]:
            cell.fill = PatternFill('solid', start_color='1F4E79')
            cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.row_dimensions[1].height = 28
        for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
            fill = PatternFill('solid', start_color='DEEAF1') if ri%2==0 else PatternFill()
            for cell in row:
                cell.fill = fill
                cell.font = Font(name='Arial', size=9)
        for col in ws.columns:
            ml = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(ml+3, 42)
        ws.freeze_panes = 'A2'
    wb.save(ruta_rep)

    nombre_archivo = f'REPORTE_MOROSOS_+3meses_{fecha_hoy}.xlsx'
    print(f'\nReporte generado: {nombre_archivo}')
    print(f'Total morosos: {len(todos)}')
    print(res.to_string(index=False))
    return ruta_rep, nombre_archivo

# MAIN
if __name__ == '__main__':
    print(f'Iniciando proceso: {datetime.datetime.now()}')

    archivos = descargar_morosos()
    print(f'\n{len(archivos)} archivos descargados')

    if not archivos:
        print('Sin archivos para analizar')
        exit(0)

    todos = analizar_morosos(archivos)
    print(f'\nTotal morosos +3 meses: {len(todos)}')

    resultado = generar_reporte(todos)
    if not resultado:
        exit(0)
    ruta_rep, nombre_archivo = resultado

    print('\nSubiendo a Google Drive...')
    service = get_drive_service()
    subir_a_drive(service, ruta_rep, nombre_archivo, FOLDER_ID)

    print(f'\nProceso completado: {datetime.datetime.now()}')
